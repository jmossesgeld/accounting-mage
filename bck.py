
import pandas as pd
import datetime
from openpyxl import load_workbook

file_path = '/'


# LOAD EXCEL FILE
wb = load_workbook(file_path)
ws = wb['MAIN']

# GET FILER DATA
RDO_CODE = ws['B1'].value
PERIOD = datetime.date.strftime(ws['B2'].value, "%m/%d/%Y")
CALENDAR = ws['B3'].value
TIN = f"{ws['B7'].value:0>9d}"
NAME = ws['B8'].value
TRADE = ws['B12'].value
ADDRESS1 = ws['B13'].value
ADDRESS2 = ws['B14'].value

# GET FILER TOTALS
EXEMPT = ws['B18'].value
ZERO_RATED = ws['B19'].value
SERVICES = ws['B20'].value
CAPITAL_GOODS = ws['B21'].value
GOODS = ws['B22'].value
INPUT_VAT = ws['B23'].value
CREDITABLE = ws['B24'].value
NON_CREDITABLE = ws['B25'].value
NO_OF_RECORDS = ws['B26'].value

# READ DATA
df = pd.read_excel(file_path, 'DATA').dropna()

# CONVERT AND FORMAT DATA
def generate_line(line):
    line[0] = f"{str(line[0]).replace('-','')}"
    line[7] = f'{line[7]:.2f}'
    line[8] = f'{line[8]:.2f}'
    line[9] = f'{line[9]:.2f}'
    line[10] = f'{line[10]:.2f}'
    line[11] = f'{line[11]:.2f}'
    line[12] = f'{line[12]:.2f}'

    for i in range(len(line)):
        try:
            line[i] = line[i].upper()
        except AttributeError:
            pass

    return f'D,P,"{line[0]}","{line[1]}",,,,"{line[5]}","{line[6]}",{line[7]},{line[8]},{line[9]},{line[10]},{line[11]},{line[12]},{TIN},{PERIOD}\n'

print(generate_line(df.values[3]))

file_types = [('BIR SLSP File', '*.DAT')]
# dest_path = filedialog.asksaveasfile(filetypes=file_types, defaultextension=file_types,
#                                         initialfile=f'{TIN}P{PERIOD[:2]}{PERIOD[6:10]}').name
dest_path = f'{TIN}P{PERIOD[:2]}{PERIOD[6:10]}.DAT'

if dest_path:
    with open(dest_path, 'w') as dat:
        dat.write('HEADER\n')
        dat.writelines([generate_line(line) for line in df.values])