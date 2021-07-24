import pandas as pd
import datetime
from openpyxl import load_workbook


def convert(file):
    # LOAD EXCEL FILE
    wb = load_workbook(file, data_only=True)
    ws = wb.worksheets[0]
    error_msgs = '<strong>Errors:</strong><br>'
    has_error = False

    # GET FILER DATA
    try:
        RDO_CODE = f"{ws['B1'].value:0>3d}"
    except Exception:
        has_error = True
        error_msgs += 'Please fill-up RDO Code<br>'

    try:
        PERIOD = datetime.date.strftime(ws['B2'].value, "%m/%d/%Y")
    except Exception:
        has_error = True
        error_msgs += 'Please fill-up PERIOD<br>'

    CALENDAR = ws['B3'].value

    try:
        TIN = f"{ws['B7'].value.replace('-','')}"
    except AttributeError:
        try:
            TIN = f"{ws['B7'].value:0>9d}"
        except Exception:
            has_error = True
            error_msgs += 'Please fill-up Filer TIN<br>'

    NAME = ws['B8'].value
    TRADE = ws['B12'].value
    ADDRESS1 = ws['B13'].value
    ADDRESS2 = ws['B14'].value

    # GET FILER TOTALS
    EXEMPT = f"{ws['B18'].value:.2f}"
    ZERO_RATED = f"{ws['B19'].value:.2f}"
    SERVICES = f"{ws['B20'].value:.2f}"
    CAPITAL_GOODS = f"{ws['B21'].value:.2f}"
    GOODS = f"{ws['B22'].value:.2f}"
    INPUT_VAT = f"{ws['B23'].value:.2f}"
    CREDITABLE = f"{ws['B24'].value:.2f}"
    NON_CREDITABLE = f"{ws['B25'].value:.2f}"
    NO_OF_RECORDS = f"{ws['B26'].value:.2f}"

    # READ AND FIX DATA
    df = pd.read_excel(file, 'DATA').fillna(0)
    df['ADDRESS 1'] = df['ADDRESS 1'].replace(0, '-')
    df['ADDRESS 2'] = df['ADDRESS 2'].replace(0, '-')
    df['LAST NAME'] = df['LAST NAME'].replace(0, '')
    df['FIRST NAME'] = df['FIRST NAME'].replace(0, '')
    df['MIDDLE NAME'] = df['MIDDLE NAME'].replace(0, '')

    # CONVERT AND FORMAT DATA
    def parse(line):
        try:
            line[0] = f"{str(line[0]).replace('-','')}"
        except AttributeError:
            line[0] = f"{str(line[0]):0>9d}"

        line[7] = f'{line[7]:.2f}'
        line[8] = f'{line[8]:.2f}'
        line[9] = f'{line[9]:.2f}'
        line[10] = f'{line[10]:.2f}'
        line[11] = f'{line[11]:.2f}'
        line[12] = f'{line[12]:.2f}'

        for i in range(len(line)):
            try:
                line[i] = line[i].upper()
            except AttributeError:
                pass

        return f'D,P,"{line[0]}","{line[1]}",,,,"{line[5]}","{line[6]}",{line[7]},{line[8]},{line[9]},{line[10]},{line[11]},{line[12]},{TIN},{PERIOD}\n'

    #RETURN RESULT
    if has_error:
        return error_msgs
    else:
        with open('result.DAT', 'w') as dat:
            dat.write(f'H,P,"{TIN}","{NAME}","","","","{TRADE}","{ADDRESS1}","{ADDRESS2}",{EXEMPT},{ZERO_RATED},{SERVICES},{CAPITAL_GOODS},{GOODS},{INPUT_VAT},{CREDITABLE},{NON_CREDITABLE},{RDO_CODE},{PERIOD},{CALENDAR}\n')
            dat.writelines([parse(line) for line in df.values])

        dest_path = f'{TIN}P{PERIOD[:2]}{PERIOD[6:10]}.DAT'
        return dest_path
