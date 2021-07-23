import os
from flask import send_from_directory
from flask import Flask, flash, request, redirect, url_for, render_template
from werkzeug.utils import secure_filename
from openpyxl import load_workbook


app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def home():
    if request.method == 'POST':
        file = request.files['file']
        # path = os.path.join("uploads", file.filename)
        path = file.filename
        print(path)
        file.save(path)
        wb = load_workbook(path)
        ws = wb['MAIN']
        print(ws['A1'].value)
        with open(file.filename,'w') as dat:
            dat.write(ws['A1'].value)

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
